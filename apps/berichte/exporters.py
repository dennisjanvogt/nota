"""
Export-Klassen für verschiedene Formate.

Alle Exporte verwenden deutsche Spaltennamen und Formatierungen.
"""
import csv
import io
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


class BaseExporter:
    """
    Basis-Klasse für alle Exporter.

    Definiert die gemeinsame Schnittstelle und Hilfsmethoden.
    """

    def __init__(self, queryset, spalten, titel='Export'):
        """
        Initialisiert den Exporter.

        Args:
            queryset: Django QuerySet mit den zu exportierenden Daten
            spalten: Liste von Tupeln (feldname, spaltenname_deutsch)
            titel: Titel für den Export
        """
        self.queryset = queryset
        self.spalten = spalten
        self.titel = titel

    def get_data_rows(self):
        """
        Extrahiert Daten-Zeilen aus dem QuerySet.

        Returns:
            List[List]: Liste von Zeilen, jede Zeile ist eine Liste von Werten
        """
        rows = []
        for obj in self.queryset:
            row = []
            for feldname, _ in self.spalten:
                # Verschachtelte Felder unterstützen (z.B. 'notarstelle__name')
                wert = obj
                for teil in feldname.split('__'):
                    wert = getattr(wert, teil, '')
                    if wert is None:
                        wert = ''

                # Formatiere Wert
                if hasattr(wert, 'strftime'):  # Datum/Zeit
                    wert = wert.strftime('%d.%m.%Y %H:%M') if hasattr(wert, 'hour') else wert.strftime('%d.%m.%Y')
                elif isinstance(wert, bool):
                    wert = 'Ja' if wert else 'Nein'
                elif hasattr(wert, 'get_display'):  # Choice-Felder
                    wert = wert.get_display()

                row.append(str(wert))
            rows.append(row)
        return rows

    def get_spalten_namen(self):
        """
        Liefert die deutschen Spaltennamen.

        Returns:
            List[str]: Liste der Spaltennamen
        """
        return [name for _, name in self.spalten]

    def export(self):
        """
        Führt den Export aus.

        Muss von Subklassen implementiert werden.

        Returns:
            HttpResponse: Response mit exportierten Daten
        """
        raise NotImplementedError("Subklassen müssen export() implementieren")


class CSVExporter(BaseExporter):
    """
    Exportiert Daten als CSV-Datei mit deutschen Spaltennamen.
    """

    def export(self):
        """
        Erstellt CSV-Export.

        Returns:
            HttpResponse: CSV-Datei zum Download
        """
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{self.titel}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        # BOM für Excel-Kompatibilität
        response.write('\ufeff')

        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_ALL)

        # Header
        writer.writerow(self.get_spalten_namen())

        # Daten
        for row in self.get_data_rows():
            writer.writerow(row)

        return response


class ExcelExporter(BaseExporter):
    """
    Exportiert Daten als Excel-Datei mit Formatierung.
    """

    def export(self):
        """
        Erstellt Excel-Export mit Formatierung.

        Returns:
            HttpResponse: Excel-Datei zum Download
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.titel[:31]  # Excel-Limit für Blattnamen

        # Header-Formatierung
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Schreibe Header
        header = self.get_spalten_namen()
        for col_num, spaltenname in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num, value=spaltenname)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Schreibe Daten
        for row_num, row_data in enumerate(self.get_data_rows(), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # Auto-Breite für Spalten
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Header-Zeile fixieren
        worksheet.freeze_panes = 'A2'

        # Speichere in BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.titel}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        return response


class PDFExporter(BaseExporter):
    """
    Exportiert Daten als PDF-Datei mit deutscher Formatierung.
    """

    def export(self):
        """
        Erstellt PDF-Export.

        Returns:
            HttpResponse: PDF-Datei zum Download
        """
        buffer = io.BytesIO()

        # Landscape für mehr Spalten
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # Container für PDF-Elemente
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#0D6EFD'),
            spaceAfter=20,
            alignment=1  # Center
        )

        # Titel
        title = Paragraph(self.titel, title_style)
        elements.append(title)

        # Datum
        datum_text = f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        datum = Paragraph(datum_text, styles['Normal'])
        elements.append(datum)
        elements.append(Spacer(1, 0.5*cm))

        # Tabelle vorbereiten
        data = [self.get_spalten_namen()]  # Header
        data.extend(self.get_data_rows())  # Daten

        # Tabelle erstellen
        table = Table(data, repeatRows=1)

        # Tabellen-Style
        table_style = TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Daten
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

            # Gitternetz
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),

            # Alternierende Zeilen
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ])

        table.setStyle(table_style)
        elements.append(table)

        # PDF generieren
        doc.build(elements)

        # Response
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.titel}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        response.write(pdf)

        return response


# Factory-Funktion für einfache Verwendung
def export_data(queryset, spalten, format_typ, titel='Export'):
    """
    Factory-Funktion zum Exportieren von Daten.

    Args:
        queryset: Django QuerySet
        spalten: Liste von (feldname, spaltenname) Tupeln
        format_typ: 'csv', 'excel' oder 'pdf'
        titel: Titel für den Export

    Returns:
        HttpResponse mit exportierten Daten
    """
    exporters = {
        'csv': CSVExporter,
        'excel': ExcelExporter,
        'pdf': PDFExporter,
    }

    exporter_class = exporters.get(format_typ.lower())
    if not exporter_class:
        return HttpResponse(
            f"Unbekannter Export-Format: {format_typ}",
            status=400
        )

    exporter = exporter_class(queryset, spalten, titel)
    return exporter.export()
