from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.utils import timezone
from apps.notarstellen.models import Notarstelle
from apps.personen.models import Notar, NotarAnwaerter
from apps.workflows.models import WorkflowTyp, WorkflowSchritt, WorkflowInstanz
from apps.berichte.exporters import CSVExporter, ExcelExporter, PDFExporter
import csv
import io
from openpyxl import load_workbook

KammerBenutzer = get_user_model()


class ExporterTestCase(TestCase):
    """Tests für Export-Funktionen."""

    def setUp(self):
        """Test-Daten erstellen."""
        # Benutzer
        self.benutzer = KammerBenutzer.objects.create_user(
            username='test',
            password='test123',
            first_name='Test',
            last_name='Benutzer',
            rolle='sachbearbeiter'
        )

        # Notarstelle
        self.notarstelle = Notarstelle.objects.create(
            notarnummer='1',
            bezeichnung='NOT-1',
            name='Notariat Hamburg I',
            strasse='Hauptstraße 1',
            plz='20095',
            stadt='Hamburg',
            bundesland='HH',
            ist_aktiv=True
        )

        # Notar
        self.notar = Notar.objects.create(
            vorname='Max',
            nachname='Mustermann',
            email='max.mustermann@example.com',
            titel='Dr.',
            notar_id='N-001',
            notarstelle=self.notarstelle,
            bestellt_am=timezone.now().date(),
            beginn_datum=timezone.now().date(),
            ist_aktiv=True
        )

        # Notariatskandidat
        self.anwaerter = NotarAnwaerter.objects.create(
            vorname='Maria',
            nachname='Musterfrau',
            email='maria.musterfrau@example.com',
            anwaerter_id='A-001',
            betreuender_notar=self.notar,
            notarstelle=self.notarstelle,
            zugelassen_am=timezone.now().date(),
            beginn_datum=timezone.now().date(),
            ist_aktiv=True
        )

        # Workflow-Typ
        self.workflow_typ = WorkflowTyp.objects.create(
            name='Testprozess',
            beschreibung='Ein Testprozess',
            ist_aktiv=True
        )

        # Workflow-Schritt
        self.workflow_schritt = WorkflowSchritt.objects.create(
            workflow_typ=self.workflow_typ,
            name='Test Schritt',
            reihenfolge=1,
            ist_optional=False
        )

        # Aktenzeichen
        self.sequenz = Nummernsequenz.objects.create(
            jahr=timezone.now().year,
            praefix='TEST',
            aktuelle_nummer=1
        )
        self.aktenzeichen = Aktenzeichen.objects.create(
            sequenz=self.sequenz,
            laufnummer=1,
            jahr=timezone.now().year,
            kategorie='allgemein',
            beschreibung='Test Aktenzeichen'
        )
        self.aktenzeichen.nummer_generieren()

        # Workflow-Instanz
        self.workflow = WorkflowInstanz.objects.create(
            workflow_typ=self.workflow_typ,
            name='Test Workflow',
            status='aktiv',
            erstellt_von=self.benutzer,
            aktenzeichen=self.aktenzeichen
        )


class CSVExporterTestCase(ExporterTestCase):
    """Tests für CSV-Export."""

    def test_notare_csv_export(self):
        """Test: Notare können als CSV exportiert werden."""
        queryset = Notar.objects.all()
        spalten = [
            ('vorname', 'Vorname'),
            ('nachname', 'Nachname'),
            ('titel', 'Titel'),
            ('notar_id', 'Notar-ID'),
            ('notarstelle__name', 'Notarstelle'),
            ('ist_aktiv', 'Aktiv')
        ]

        exporter = CSVExporter(queryset, spalten)
        response = exporter.export()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv; charset=utf-8')
        self.assertIn('attachment; filename="', response['Content-Disposition'])
        self.assertIn('.csv"', response['Content-Disposition'])

        # CSV-Inhalt prüfen
        content = response.content.decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content), delimiter=';')
        rows = list(reader)

        # Header prüfen
        self.assertEqual(rows[0], ['Vorname', 'Nachname', 'Titel', 'Notar-ID', 'Notarstelle', 'Aktiv'])

        # Daten prüfen
        self.assertEqual(rows[1][0], 'Max')
        self.assertEqual(rows[1][1], 'Mustermann')
        self.assertEqual(rows[1][2], 'Dr.')
        self.assertEqual(rows[1][3], 'N-001')
        self.assertEqual(rows[1][4], 'Notariat Hamburg I')
        self.assertEqual(rows[1][5], 'Ja')  # Boolean als "Ja"

    def test_anwaerter_csv_export(self):
        """Test: Notariatskandidat können als CSV exportiert werden."""
        queryset = NotarAnwaerter.objects.all()
        spalten = [
            ('vorname', 'Vorname'),
            ('nachname', 'Nachname'),
            ('anwaerter_id', 'Kandidaten-ID'),
            ('betreuender_notar__nachname', 'Betreuender Notar')
        ]

        exporter = CSVExporter(queryset, spalten)
        response = exporter.export()

        self.assertEqual(response.status_code, 200)

        content = response.content.decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content), delimiter=';')
        rows = list(reader)

        self.assertEqual(rows[1][0], 'Maria')
        self.assertEqual(rows[1][1], 'Musterfrau')
        self.assertEqual(rows[1][2], 'A-001')
        self.assertEqual(rows[1][3], 'Mustermann')


class ExcelExporterTestCase(ExporterTestCase):
    """Tests für Excel-Export."""

    def test_notare_excel_export(self):
        """Test: Notare können als Excel exportiert werden."""
        queryset = Notar.objects.all()
        spalten = [
            ('vorname', 'Vorname'),
            ('nachname', 'Nachname'),
            ('notar_id', 'Notar-ID'),
            ('ist_aktiv', 'Aktiv')
        ]

        exporter = ExcelExporter(queryset, spalten, 'Notare')
        response = exporter.export()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment; filename="', response['Content-Disposition'])
        self.assertIn('.xlsx"', response['Content-Disposition'])

        # Excel-Inhalt prüfen
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Header prüfen (fett)
        self.assertEqual(ws.cell(1, 1).value, 'Vorname')
        self.assertEqual(ws.cell(1, 2).value, 'Nachname')
        self.assertEqual(ws.cell(1, 3).value, 'Notar-ID')
        self.assertEqual(ws.cell(1, 4).value, 'Aktiv')
        self.assertTrue(ws.cell(1, 1).font.bold)

        # Daten prüfen
        self.assertEqual(ws.cell(2, 1).value, 'Max')
        self.assertEqual(ws.cell(2, 2).value, 'Mustermann')
        self.assertEqual(ws.cell(2, 3).value, 'N-001')
        self.assertEqual(ws.cell(2, 4).value, 'Ja')

    def test_notarstellen_excel_export(self):
        """Test: Notarstellen können als Excel exportiert werden."""
        queryset = Notarstelle.objects.all()
        spalten = [
            ('notarnummer', 'Notarnummer'),
            ('name', 'Name'),
            ('stadt', 'Stadt'),
            ('ist_aktiv', 'Aktiv')
        ]

        exporter = ExcelExporter(queryset, spalten, 'Notarstellen')
        response = exporter.export()

        self.assertEqual(response.status_code, 200)

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        self.assertEqual(ws.cell(2, 1).value, '1')
        self.assertEqual(ws.cell(2, 2).value, 'Notariat Hamburg I')
        self.assertEqual(ws.cell(2, 3).value, 'Hamburg')
        self.assertEqual(ws.cell(2, 4).value, 'Ja')


class PDFExporterTestCase(ExporterTestCase):
    """Tests für PDF-Export."""

    def test_notare_pdf_export(self):
        """Test: Notare können als PDF exportiert werden."""
        queryset = Notar.objects.all()
        spalten = [
            ('vorname', 'Vorname'),
            ('nachname', 'Nachname'),
            ('notar_id', 'Notar-ID')
        ]

        exporter = PDFExporter(queryset, spalten, 'Notare')
        response = exporter.export()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('attachment; filename="', response['Content-Disposition'])
        self.assertIn('.pdf"', response['Content-Disposition'])

        # PDF sollte Inhalt haben
        self.assertGreater(len(response.content), 0)

    def test_workflows_pdf_export(self):
        """Test: Workflows können als PDF exportiert werden."""
        queryset = WorkflowInstanz.objects.all()
        spalten = [
            ('name', 'Name'),
            ('status', 'Status'),
            ('aktenzeichen__vollstaendige_nummer', 'Aktenzeichen')
        ]

        exporter = PDFExporter(queryset, spalten, 'Workflows')
        response = exporter.export()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertGreater(len(response.content), 0)


class BerichteViewsTestCase(ExporterTestCase):
    """Tests für Berichte-Views."""

    def setUp(self):
        super().setUp()
        self.client = Client()
        self.client.login(username='test', password='test123')

    def test_uebersicht_view(self):
        """Test: Berichte-Übersicht ist erreichbar."""
        response = self.client.get(reverse('berichte_uebersicht'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'berichte/uebersicht.html')

    def test_notare_export_csv(self):
        """Test: Notare CSV-Export funktioniert."""
        response = self.client.get(reverse('export_notare') + '?format=csv')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv; charset=utf-8')

    def test_notare_export_excel(self):
        """Test: Notare Excel-Export funktioniert."""
        response = self.client.get(reverse('export_notare') + '?format=excel')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def test_notare_export_pdf(self):
        """Test: Notare PDF-Export funktioniert."""
        response = self.client.get(reverse('export_notare') + '?format=pdf')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_anwaerter_export_csv(self):
        """Test: Kandidat CSV-Export funktioniert."""
        response = self.client.get(reverse('export_anwaerter') + '?format=csv')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv; charset=utf-8')

    def test_notarstellen_export_excel(self):
        """Test: Notarstellen Excel-Export funktioniert."""
        response = self.client.get(reverse('export_notarstellen') + '?format=excel')
        self.assertEqual(response.status_code, 200)

    def test_workflows_export_csv(self):
        """Test: Workflows CSV-Export funktioniert."""
        response = self.client.get(reverse('export_workflows') + '?format=csv')
        self.assertEqual(response.status_code, 200)

    def test_aktenzeichen_export_pdf(self):
        """Test: Aktenzeichen PDF-Export funktioniert."""
        response = self.client.get(reverse('export_aktenzeichen') + '?format=pdf')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_invalid_format(self):
        """Test: Ungültiges Format gibt 400 zurück."""
        response = self.client.get(reverse('export_notare') + '?format=invalid')
        self.assertEqual(response.status_code, 400)

    def test_login_required(self):
        """Test: Login ist erforderlich für Berichte."""
        self.client.logout()
        response = self.client.get(reverse('berichte_uebersicht'))
        self.assertEqual(response.status_code, 302)  # Redirect to login
        self.assertIn('/login/', response.url)


class DateFormatTestCase(ExporterTestCase):
    """Tests für Datumsformatierung in Exporten."""

    def test_date_format_in_csv(self):
        """Test: Datumswerte werden korrekt formatiert."""
        queryset = Notar.objects.all()
        spalten = [
            ('nachname', 'Nachname'),
            ('bestellt_am', 'Bestellt am')
        ]

        exporter = CSVExporter(queryset, spalten)
        response = exporter.export()

        content = response.content.decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content), delimiter=';')
        rows = list(reader)

        # Datum sollte im Format DD.MM.YYYY sein
        datum = rows[1][1]
        self.assertRegex(datum, r'\d{2}\.\d{2}\.\d{4}')

    def test_none_value_handling(self):
        """Test: None-Werte werden als leerer String exportiert."""
        # Notar ohne Titel
        notar_ohne_titel = Notar.objects.create(
            vorname='Hans',
            nachname='Schmidt',
            email='hans.schmidt@example.com',  # Eindeutige E-Mail
            notar_id='N-002',
            notarstelle=self.notarstelle,
            bestellt_am=timezone.now().date(),
            beginn_datum=timezone.now().date(),
            ist_aktiv=True
            # titel ist None
        )

        queryset = Notar.objects.filter(id=notar_ohne_titel.id)
        spalten = [
            ('nachname', 'Nachname'),
            ('titel', 'Titel')
        ]

        exporter = CSVExporter(queryset, spalten)
        response = exporter.export()

        content = response.content.decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content), delimiter=';')
        rows = list(reader)

        # Titel sollte leer sein, nicht "None"
        self.assertEqual(rows[1][1], '')


class GermanColumnNamesTestCase(ExporterTestCase):
    """Tests für deutsche Spaltennamen."""

    def test_german_column_names_in_csv(self):
        """Test: Deutsche Spaltennamen werden verwendet."""
        queryset = Notar.objects.all()
        spalten = [
            ('vorname', 'Vorname'),
            ('nachname', 'Nachname')
        ]

        exporter = CSVExporter(queryset, spalten)
        response = exporter.export()

        content = response.content.decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content), delimiter=';')
        rows = list(reader)

        # Header sollte deutsche Namen haben
        self.assertEqual(rows[0][0], 'Vorname')
        self.assertEqual(rows[0][1], 'Nachname')

    def test_german_column_names_in_excel(self):
        """Test: Deutsche Spaltennamen in Excel."""
        queryset = Notarstelle.objects.all()
        spalten = [
            ('notarnummer', 'Notarnummer'),
            ('name', 'Name'),
            ('stadt', 'Stadt')
        ]

        exporter = ExcelExporter(queryset, spalten, 'Notarstellen')
        response = exporter.export()

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        self.assertEqual(ws.cell(1, 1).value, 'Notarnummer')
        self.assertEqual(ws.cell(1, 2).value, 'Name')
        self.assertEqual(ws.cell(1, 3).value, 'Stadt')
